import openpyxl

path="..//ExcelFiles//writingTestData.xlsx"
sheetName="Writing"

def getRowCount(path, sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColumnCount(path, sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_column


def getCellData(path, sheetName, rowNum, colNum):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNum, column=colNum).value


def setCellData(path, sheetName, rowNum, colNum, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNum, column=colNum).value = data
    workbook.save(path)
rows=getRowCount(path,sheetName)
cols=getColumnCount(path,sheetName)

print(rows,cols)
print(getCellData(path,sheetName,1,4))
setCellData(path,sheetName,1,4,"New value 1")
print(getCellData(path,sheetName,1,4))