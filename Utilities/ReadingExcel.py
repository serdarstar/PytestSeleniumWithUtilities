import openpyxl

workbook = openpyxl.load_workbook("..//ExcelFiles//testdata.xlsx")
sheet = workbook["LoginTest"]
totalRows=sheet.max_row
totalColums=sheet.max_column

print("Total rows is: ", totalRows, "Total columns is: ",totalColums)

for row in range(1,totalRows+1):
    for column in range(1, totalColums+1):
        print(sheet.cell(row=row, column=column).value, end=" ")
    print()